"""Unit tests for the merged Excel export feature.

Tests edge cases and formatting details for the merged calendar worksheet:
- Time-range font size ≤ 8pt
- Backward-compatible export() signature
- OSError with descriptive message on bad filepath
- Column widths ≥ 18 characters
- export() without shift_windows produces valid workbook
- "nobody" slots render only marker (and optionally time), no names

Requirements: 2.4, 3.3, 4.2, 4.3, 4.4
"""

import inspect
import tempfile
import os
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from dc_shiftmaster.excel_export import ExcelExporter
from dc_shiftmaster.models import ScheduleSlot, ShiftWindow
from dc_shiftmaster.scheduling import SchedulingEngine


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _minimal_schedule(year: int = 2025) -> list[ScheduleSlot]:
    """Build a minimal schedule with a few days for testing."""
    engine = SchedulingEngine()
    slots = []
    # Generate a week of slots starting Jan 1
    for day_offset in range(7):
        d = date(year, 1, 1 + day_offset)
        owner = engine.get_day_owner(d)
        day_names = ["Alice", "Bob"] if owner == "back_half" else ["Carol", "Dave"]
        night_names = ["Eve", "Frank"] if owner == "back_half" else ["Grace", "Hank"]
        slots.append(ScheduleSlot(
            date=d, shift_type="day", start_time="06:00",
            teammates=day_names, is_override=False,
        ))
        slots.append(ScheduleSlot(
            date=d, shift_type="night", start_time="18:00",
            teammates=night_names, is_override=False,
        ))
    return slots


def _shift_windows() -> dict[str, ShiftWindow]:
    """Standard day/night shift windows for testing."""
    return {
        "day": ShiftWindow(shift_type="day", start_time="06:00", end_time="18:30"),
        "night": ShiftWindow(shift_type="night", start_time="18:00", end_time="06:30"),
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestTimeRangeFontSize:
    """Requirement 2.4: Time-range text uses ≤ 8pt font size."""

    def test_marker_cells_use_8pt_or_smaller_font(self):
        """Marker cells in the merged sheet should use font size ≤ 8."""
        year = 2025
        engine = SchedulingEngine()
        schedule = _minimal_schedule(year)
        shift_windows = _shift_windows()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name

        try:
            exporter = ExcelExporter()
            exporter.export(year, schedule, engine, filepath, shift_windows=shift_windows)

            wb = load_workbook(filepath)
            ws = wb[f"{year} Shift Calendar"]

            # Find marker cells (cells containing "B\n" or "F\n" indicating multi-line content)
            found_marker = False
            for row in ws.iter_rows(min_row=3):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Marker cells start with "B" or "F" and have time info
                        if cell.value.startswith(("B\n", "F\n")):
                            found_marker = True
                            assert cell.font.size is not None
                            assert cell.font.size <= 8, (
                                f"Font size {cell.font.size} > 8 at cell {cell.coordinate}"
                            )

            assert found_marker, "No marker cells with time ranges found in the worksheet"
        finally:
            os.unlink(filepath)


class TestExportSignature:
    """Requirement 4.2: export() accepts the documented parameters (backward-compatible)."""

    def test_export_signature_has_expected_parameters(self):
        """export() must accept (self, year, schedule, engine, filepath, shift_windows=None)."""
        sig = inspect.signature(ExcelExporter.export)
        params = list(sig.parameters.keys())

        assert "self" in params
        assert "year" in params
        assert "schedule" in params
        assert "engine" in params
        assert "filepath" in params
        assert "shift_windows" in params

        # shift_windows should have a default of None
        shift_windows_param = sig.parameters["shift_windows"]
        assert shift_windows_param.default is None


class TestOSErrorOnBadFilepath:
    """Requirement 4.4: OSError with descriptive message on bad filepath."""

    def test_raises_oserror_with_descriptive_message(self):
        """export() should raise OSError with the filepath in the message."""
        year = 2025
        engine = SchedulingEngine()
        schedule = _minimal_schedule(year)
        bad_path = "/nonexistent/directory/that/does/not/exist/output.xlsx"

        exporter = ExcelExporter()
        with pytest.raises(OSError) as exc_info:
            exporter.export(year, schedule, engine, bad_path)

        error_msg = str(exc_info.value)
        assert bad_path in error_msg, (
            f"Error message should contain the filepath. Got: {error_msg}"
        )
        assert "Cannot save Excel to" in error_msg


class TestColumnWidths:
    """Requirement 3.3: Column widths are ≥ 18 characters."""

    def test_merged_sheet_column_widths_at_least_18(self):
        """All data columns in the merged calendar should be ≥ 18 wide."""
        year = 2025
        engine = SchedulingEngine()
        schedule = _minimal_schedule(year)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name

        try:
            exporter = ExcelExporter()
            exporter.export(year, schedule, engine, filepath, shift_windows=_shift_windows())

            wb = load_workbook(filepath)
            ws = wb[f"{year} Shift Calendar"]

            # Columns 1-8 should all be ≥ 18
            from openpyxl.utils import get_column_letter
            for col in range(1, 9):
                col_letter = get_column_letter(col)
                width = ws.column_dimensions[col_letter].width
                assert width >= 18, (
                    f"Column {col_letter} width is {width}, expected ≥ 18"
                )
        finally:
            os.unlink(filepath)


class TestExportWithoutShiftWindows:
    """Requirement 4.3: Calling export() without shift_windows still produces valid workbook."""

    def test_export_without_shift_windows_produces_valid_workbook(self):
        """export() with shift_windows=None should still produce a valid 3-sheet workbook."""
        year = 2025
        engine = SchedulingEngine()
        schedule = _minimal_schedule(year)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            filepath = f.name

        try:
            exporter = ExcelExporter()
            # Call without shift_windows (default None)
            exporter.export(year, schedule, engine, filepath)

            wb = load_workbook(filepath)
            # Should have 3 sheets
            assert len(wb.sheetnames) == 3
            assert wb.sheetnames[0] == f"{year} Shift Calendar"
            assert wb.sheetnames[1] == "Coverage Summary"
            assert wb.sheetnames[2] == "Timeline"

            # Calendar sheet should have content
            ws = wb[f"{year} Shift Calendar"]
            assert ws["A1"].value == f"{year} Shift Calendar"
        finally:
            os.unlink(filepath)


class TestNobodySlotsRenderOnlyMarker:
    """'nobody' slots render only marker (and optionally time), no names."""

    def test_format_cell_content_nobody_omits_names(self):
        """_format_cell_content with nobody slot should not include names."""
        exporter = ExcelExporter()
        shift_window = ShiftWindow(shift_type="day", start_time="06:00", end_time="18:30")
        nobody_slot = ScheduleSlot(
            date=date(2025, 1, 1), shift_type="day", start_time="06:00",
            teammates=["nobody"], is_override=False,
        )

        # With shift_window: should be "B\n06:00–18:30" (no names)
        content = exporter._format_cell_content("B", shift_window, nobody_slot)
        lines = content.split("\n")
        assert lines[0] == "B"
        assert lines[1] == "06:00\u201318:30"
        assert len(lines) == 2, f"Expected 2 lines (marker + time), got {len(lines)}: {lines}"

    def test_format_cell_content_nobody_without_shift_window(self):
        """_format_cell_content with nobody slot and no shift_window: marker only."""
        exporter = ExcelExporter()
        nobody_slot = ScheduleSlot(
            date=date(2025, 1, 1), shift_type="day", start_time="06:00",
            teammates=["nobody"], is_override=False,
        )

        # Without shift_window: should be just "F"
        content = exporter._format_cell_content("F", None, nobody_slot)
        assert content == "F"
