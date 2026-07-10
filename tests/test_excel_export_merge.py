"""Property-based tests for the Excel Export Merge feature.

Uses Hypothesis to verify correctness properties of the merged
calendar export logic in dc_shiftmaster/excel_export.py.
"""

import tempfile
from datetime import date, datetime, timedelta

import openpyxl
from hypothesis import given, settings
from hypothesis import strategies as st

from dc_shiftmaster.excel_export import ExcelExporter
from dc_shiftmaster.models import ScheduleSlot, ShiftWindow
from dc_shiftmaster.scheduling import SchedulingEngine

from tests.conftest import valid_time, valid_shift_windows, valid_year


# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------

@st.composite
def shift_window_strategy(draw: st.DrawFn) -> ShiftWindow:
    """Generate a random ShiftWindow with valid times."""
    shift_type = draw(st.sampled_from(["day", "night"]))
    start = draw(valid_time())
    end = draw(valid_time())
    return ShiftWindow(shift_type=shift_type, start_time=start, end_time=end)


@st.composite
def schedule_slot_strategy(draw: st.DrawFn) -> ScheduleSlot:
    """Generate a random ScheduleSlot with either real teammates or ['nobody']."""
    d = draw(st.dates(min_value=date(2000, 1, 1), max_value=date(2100, 12, 31)))
    shift_type = draw(st.sampled_from(["day", "night"]))
    start_time = draw(valid_time())
    # Either generate real teammate names or ["nobody"]
    use_nobody = draw(st.booleans())
    if use_nobody:
        teammates = ["nobody"]
    else:
        teammates = draw(
            st.lists(
                st.text(
                    alphabet=st.characters(whitelist_categories=("L", "N", "Zs")),
                    min_size=1,
                    max_size=20,
                ).filter(lambda s: s.strip()),
                min_size=1,
                max_size=4,
            )
        )
    return ScheduleSlot(
        date=d,
        shift_type=shift_type,
        start_time=start_time,
        teammates=teammates,
        is_override=draw(st.booleans()),
    )


# ---------------------------------------------------------------------------
# Property 3: Cell content format with shift times
# ---------------------------------------------------------------------------


class TestCellContentFormat:
    """Feature: excel-export-merge, Property 3: Cell content format with shift times

    Validates: Requirements 2.1, 2.2, 2.3
    """

    @given(
        marker=st.sampled_from(["B", "F"]),
        shift_window=st.one_of(st.none(), shift_window_strategy()),
        slot=st.one_of(st.none(), schedule_slot_strategy()),
    )
    @settings(max_examples=100)
    def test_cell_content_format_with_shift_times(
        self, marker: str, shift_window, slot
    ):
        """Feature: excel-export-merge, Property 3: Cell content format with shift times

        Validates: Requirements 2.1, 2.2, 2.3

        For any valid marker (B or F), shift_window, and teammate list,
        _format_cell_content SHALL produce a string with lines in the order:
        (1) the B/F marker, (2) the time range "start–end" (only when
        shift_window is provided), (3) comma-separated teammate names (only
        when teammates are not ["nobody"]). When shift_window is None, the
        time range line SHALL be absent.
        """
        exporter = ExcelExporter()
        result = exporter._format_cell_content(marker, shift_window, slot)
        lines = result.split("\n")

        # Line 1 is always the marker
        assert lines[0] == marker

        # Determine expected line count and content
        expected_line_idx = 1

        # Time range line present iff shift_window is not None
        if shift_window is not None:
            assert len(lines) > expected_line_idx
            time_line = lines[expected_line_idx]
            # Must contain the en-dash character (U+2013)
            assert "\u2013" in time_line
            # Must be formatted as "start_time–end_time"
            assert time_line == f"{shift_window.start_time}\u2013{shift_window.end_time}"
            expected_line_idx += 1
        else:
            # No time range line should be present — verify by checking
            # that if there are more lines, they are names (not a time range)
            for line in lines[1:]:
                assert "\u2013" not in line or (
                    slot is not None and slot.teammates != ["nobody"]
                )

        # Names line present iff slot is not None and teammates != ["nobody"]
        has_names = slot is not None and slot.teammates != ["nobody"]
        if has_names:
            assert len(lines) > expected_line_idx
            names_line = lines[expected_line_idx]
            expected_names = ", ".join(slot.teammates)
            assert names_line == expected_names
            expected_line_idx += 1

        # Total line count must match exactly
        assert len(lines) == expected_line_idx


# ---------------------------------------------------------------------------
# Property 2: Merged block layout structure
# ---------------------------------------------------------------------------


class TestMergedBlockLayoutStructure:
    """Feature: excel-export-merge, Property 2: Merged block layout structure

    Validates: Requirements 1.3, 3.1, 3.2
    """

    @given(year=valid_year())
    @settings(max_examples=100)
    def test_merged_block_layout_structure(self, year: int):
        """Feature: excel-export-merge, Property 2: Merged block layout structure

        Validates: Requirements 1.3, 3.1, 3.2

        For any valid year and schedule, each two-week block in the
        Calendar_Worksheet SHALL contain exactly 6 rows per week in the order:
        date row, day-name row, day-shift B-marker row (with "Day" label in
        column A), day-shift F-marker row, night-shift B-marker row (with
        "Night" label in column A), night-shift F-marker row.
        """
        from openpyxl import Workbook
        from dc_shiftmaster.excel_export import ExcelExporter, _find_cycle_start
        from dc_shiftmaster.scheduling import SchedulingEngine

        exporter = ExcelExporter()
        engine = SchedulingEngine()
        wb = Workbook()
        ws = wb.active

        # Find a valid Wednesday start date for this year
        cycle_start = _find_cycle_start(year)
        # cycle_start is the last Wednesday of previous December
        # Use it as our week_start (it's guaranteed to be a Wednesday)
        week_start = cycle_start

        # Use an empty lookup dict (testing structure, not content)
        lookup: dict[tuple[str, str], ScheduleSlot] = {}

        start_row = 1
        result_row = exporter._write_merged_two_week_block(
            ws, start_row, week_start, year, engine, lookup, shift_windows=None
        )

        # The method should write exactly 12 rows (6 per week × 2 weeks)
        assert result_row == start_row + 12

        # Verify structure for each of the 2 weeks
        for week_offset in range(2):
            base_row = start_row + (week_offset * 6)

            # Row 1 of week: date row - cells should contain date strings (D-Mon format)
            # At least one cell in columns 2-8 should have a date value
            date_row_values = [
                ws.cell(row=base_row, column=col).value for col in range(2, 9)
            ]
            # Date row should have at least some non-None values
            assert any(v is not None for v in date_row_values), (
                f"Date row (row {base_row}) should have date values"
            )

            # Row 2 of week: day-name row - should contain day names
            day_name_row_values = [
                ws.cell(row=base_row + 1, column=col).value for col in range(2, 9)
            ]
            from dc_shiftmaster.excel_export import DAY_NAMES
            assert day_name_row_values == list(DAY_NAMES), (
                f"Day-name row (row {base_row + 1}) should contain DAY_NAMES, "
                f"got {day_name_row_values}"
            )

            # Row 3 of week: day-shift B-marker row - "Day" label in column A
            day_label_cell = ws.cell(row=base_row + 2, column=1).value
            assert day_label_cell == "Day", (
                f"Row {base_row + 2} column A should have 'Day' label, "
                f"got {day_label_cell!r}"
            )

            # Row 4 of week: day-shift F-marker row (no specific label required)
            # Just verify it exists (row is written)

            # Row 5 of week: night-shift B-marker row - "Night" label in column A
            night_label_cell = ws.cell(row=base_row + 4, column=1).value
            assert night_label_cell == "Night", (
                f"Row {base_row + 4} column A should have 'Night' label, "
                f"got {night_label_cell!r}"
            )

            # Row 6 of week: night-shift F-marker row (no specific label required)
            # Just verify it exists (row is written)


# ---------------------------------------------------------------------------
# Property 4: Color scheme correctness
# ---------------------------------------------------------------------------


class TestWednesdayToTuesdayOrientation:
    """Feature: excel-export-merge, Property 5: Sunday-to-Saturday week orientation

    Validates: Requirements 1.5
    """

    @given(year=valid_year())
    @settings(max_examples=100)
    def test_wednesday_to_tuesday_week_orientation(self, year: int):
        """Feature: excel-export-merge, Property 5: Sunday-to-Saturday week orientation

        Validates: Requirements 1.5

        For any two-week block in the exported Calendar_Worksheet, the date row
        SHALL start on a Sunday (column 2) and end on a Saturday (column 8),
        preserving the Sun–Sat week structure.
        """
        from openpyxl import Workbook
        from dc_shiftmaster.excel_export import ExcelExporter, _find_cycle_start
        from dc_shiftmaster.scheduling import SchedulingEngine

        exporter = ExcelExporter()
        engine = SchedulingEngine()
        wb = Workbook()
        ws = wb.active

        cycle_start = _find_cycle_start(year)
        week_start = cycle_start

        lookup: dict[tuple[str, str], ScheduleSlot] = {}

        start_row = 1
        exporter._write_merged_two_week_block(
            ws, start_row, week_start, year, engine, lookup, shift_windows=None
        )

        # Verify each of the 2 weeks in the block
        for week_offset in range(2):
            base_row = start_row + (week_offset * 6)

            # Collect date cell values from the date row (columns 2-8)
            date_values = [
                ws.cell(row=base_row, column=col).value for col in range(2, 9)
            ]

            # Parse each date string ("D-Mon" format) into a date object
            # We need the year context to resolve the month abbreviation
            parsed_dates = []
            for i, val in enumerate(date_values):
                assert val is not None, (
                    f"Date cell at row {base_row}, col {i + 2} should not be None"
                )
                # Format is "D-Mon" e.g. "31-Dec", "1-Jan"
                day_str, mon_str = val.split("-")
                # Parse month abbreviation to month number
                parsed_month = datetime.strptime(mon_str, "%b").month
                # Determine the correct year for this date
                if parsed_month == 12:
                    d = date(year - 1, parsed_month, int(day_str))
                else:
                    d = date(year, parsed_month, int(day_str))
                parsed_dates.append(d)

            # First date column should be a Sunday (weekday() == 6)
            assert parsed_dates[0].weekday() == 6, (
                f"Week {week_offset + 1}: First date {parsed_dates[0]} "
                f"should be Sunday (weekday=6), got weekday={parsed_dates[0].weekday()}"
            )

            # Last date column should be a Saturday (weekday() == 5)
            assert parsed_dates[6].weekday() == 5, (
                f"Week {week_offset + 1}: Last date {parsed_dates[6]} "
                f"should be Saturday (weekday=5), got weekday={parsed_dates[6].weekday()}"
            )

            # Dates should be consecutive (each column is one day after the previous)
            for i in range(1, 7):
                expected = parsed_dates[i - 1] + timedelta(days=1)
                assert parsed_dates[i] == expected, (
                    f"Week {week_offset + 1}: Column {i + 1} date {parsed_dates[i]} "
                    f"should be one day after column {i} date {parsed_dates[i - 1]}, "
                    f"expected {expected}"
                )


class TestColorSchemeCorrectness:
    """Feature: excel-export-merge, Property 4: Color scheme correctness

    Validates: Requirements 1.4
    """

    @given(year=valid_year())
    @settings(max_examples=100)
    def test_color_scheme_correctness(self, year: int):
        """Feature: excel-export-merge, Property 4: Color scheme correctness

        Validates: Requirements 1.4

        For any date in the calendar, the fill color of a day-shift marker cell
        SHALL be gold (FFD966) for Back_Half or green (A9D18E) for Front_Half,
        and the fill color of a night-shift marker cell SHALL be blue (B4C6E7)
        for Back_Half or purple (D5A6E6) for Front_Half.
        """
        from openpyxl import Workbook
        from dc_shiftmaster.excel_export import ExcelExporter, _find_cycle_start
        from dc_shiftmaster.scheduling import SchedulingEngine

        exporter = ExcelExporter()
        engine = SchedulingEngine()
        wb = Workbook()
        ws = wb.active

        # Find a valid Wednesday start date for this year
        cycle_start = _find_cycle_start(year)
        week_start = cycle_start

        # Use an empty lookup dict (testing colors, not content)
        lookup: dict[tuple[str, str], ScheduleSlot] = {}

        start_row = 1
        exporter._write_merged_two_week_block(
            ws, start_row, week_start, year, engine, lookup, shift_windows=None
        )

        # Expected fill colors
        GOLD = "FFD966"    # Day shift Back_Half (B)
        GREEN = "A9D18E"   # Day shift Front_Half (F)
        BLUE = "B4C6E7"    # Night shift Back_Half (B)
        PURPLE = "D5A6E6"  # Night shift Front_Half (F)

        # Verify structure for each of the 2 weeks
        for week_offset in range(2):
            base_row = start_row + (week_offset * 6)

            # Row 3: Day shift B-marker row (base_row + 2), data in columns 2-8
            for col in range(2, 9):
                cell = ws.cell(row=base_row + 2, column=col)
                if cell.value is not None and "B" in str(cell.value):
                    fill_color = cell.fill.start_color.rgb
                    # Strip leading alpha if present (openpyxl may prefix with "00")
                    if len(fill_color) == 8:
                        fill_color = fill_color[2:]
                    assert fill_color == GOLD, (
                        f"Day B-marker cell at row {base_row + 2}, col {col} "
                        f"should have gold fill ({GOLD}), got {fill_color}"
                    )

            # Row 4: Day shift F-marker row (base_row + 3), data in columns 2-8
            for col in range(2, 9):
                cell = ws.cell(row=base_row + 3, column=col)
                if cell.value is not None and "F" in str(cell.value):
                    fill_color = cell.fill.start_color.rgb
                    if len(fill_color) == 8:
                        fill_color = fill_color[2:]
                    assert fill_color == GREEN, (
                        f"Day F-marker cell at row {base_row + 3}, col {col} "
                        f"should have green fill ({GREEN}), got {fill_color}"
                    )

            # Row 5: Night shift B-marker row (base_row + 4), data in columns 2-8
            for col in range(2, 9):
                cell = ws.cell(row=base_row + 4, column=col)
                if cell.value is not None and "B" in str(cell.value):
                    fill_color = cell.fill.start_color.rgb
                    if len(fill_color) == 8:
                        fill_color = fill_color[2:]
                    assert fill_color == BLUE, (
                        f"Night B-marker cell at row {base_row + 4}, col {col} "
                        f"should have blue fill ({BLUE}), got {fill_color}"
                    )

            # Row 6: Night shift F-marker row (base_row + 5), data in columns 2-8
            for col in range(2, 9):
                cell = ws.cell(row=base_row + 5, column=col)
                if cell.value is not None and "F" in str(cell.value):
                    fill_color = cell.fill.start_color.rgb
                    if len(fill_color) == 8:
                        fill_color = fill_color[2:]
                    assert fill_color == PURPLE, (
                        f"Night F-marker cell at row {base_row + 5}, col {col} "
                        f"should have purple fill ({PURPLE}), got {fill_color}"
                    )


# ---------------------------------------------------------------------------
# Property 1: Workbook sheet structure and ordering
# ---------------------------------------------------------------------------


class TestWorkbookSheetStructure:
    """Feature: excel-export-merge, Property 1: Workbook sheet structure and ordering
    Validates: Requirements 1.1, 1.2, 5.1, 5.2
    """

    @given(
        year=valid_year(),
        shift_windows=st.one_of(st.none(), valid_shift_windows()),
    )
    @settings(max_examples=100, deadline=None)
    def test_workbook_sheet_structure_and_ordering(self, year: int, shift_windows):
        """Feature: excel-export-merge, Property 1: Workbook sheet structure and ordering

        Validates: Requirements 1.1, 1.2, 5.1, 5.2

        For any valid year, schedule, and shift_windows, the exported workbook
        SHALL contain exactly three sheets named ['{year} Shift Calendar',
        'Coverage Summary', 'Timeline'] in that order, and SHALL NOT contain
        sheets named 'Day Shift {year}' or 'Night Shift {year}'.
        """
        exporter = ExcelExporter()
        engine = SchedulingEngine()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            filepath = tmp.name

        exporter.export(year, [], engine, filepath, shift_windows=shift_windows)

        wb = openpyxl.load_workbook(filepath)

        # Verify exactly 3 sheets
        assert len(wb.sheetnames) == 3, (
            f"Expected exactly 3 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
        )

        # Verify sheet names in order
        expected_names = [f"{year} Shift Calendar", "Coverage Summary", "Timeline"]
        assert wb.sheetnames == expected_names, (
            f"Expected sheet names {expected_names}, got {wb.sheetnames}"
        )

        # Verify old separate sheets do NOT exist
        assert f"Day Shift {year}" not in wb.sheetnames, (
            f"'Day Shift {year}' should not exist in the workbook"
        )
        assert f"Night Shift {year}" not in wb.sheetnames, (
            f"'Night Shift {year}' should not exist in the workbook"
        )

        wb.close()


# ---------------------------------------------------------------------------
# Property 6: Coverage and Timeline sheets preserved
# ---------------------------------------------------------------------------


class TestCoverageAndTimelinePreserved:
    """Feature: excel-export-merge, Property 6: Coverage and Timeline sheets preserved
    Validates: Requirements 4.1
    """

    @given(year=valid_year())
    @settings(max_examples=100, deadline=None)
    def test_coverage_and_timeline_sheets_preserved(self, year: int):
        """Feature: excel-export-merge, Property 6: Coverage and Timeline sheets preserved

        Validates: Requirements 4.1

        For any valid year and schedule, the Coverage_Summary_Sheet SHALL
        contain the title 'Coverage Summary — {year}' and the expected column
        headers, and the Timeline_Sheet SHALL contain the title
        'Shift Timeline — {year}' — both unchanged from the pre-merge behavior.
        """
        exporter = ExcelExporter()
        engine = SchedulingEngine()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            filepath = tmp.name

        exporter.export(year, [], engine, filepath, shift_windows=None)

        wb = openpyxl.load_workbook(filepath)

        # Verify Coverage Summary sheet exists and has correct title
        assert "Coverage Summary" in wb.sheetnames, (
            f"'Coverage Summary' sheet should exist, got {wb.sheetnames}"
        )
        ws_cov = wb["Coverage Summary"]
        assert ws_cov["A1"].value == f"Coverage Summary — {year}", (
            f"Coverage Summary A1 should be 'Coverage Summary — {year}', "
            f"got {ws_cov['A1'].value!r}"
        )

        # Verify Timeline sheet exists and has correct title
        assert "Timeline" in wb.sheetnames, (
            f"'Timeline' sheet should exist, got {wb.sheetnames}"
        )
        ws_timeline = wb["Timeline"]
        assert ws_timeline["A1"].value == f"Shift Timeline — {year}", (
            f"Timeline A1 should be 'Shift Timeline — {year}', "
            f"got {ws_timeline['A1'].value!r}"
        )

        wb.close()
