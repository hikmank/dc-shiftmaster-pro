"""Excel exporter for DC-ShiftMaster Pro.

Generates an .xlsx file matching the ATL Back-Front Shift Rotation
spreadsheet format: a calendar grid with Wed–Tue weeks showing "B"
and "F" markers for Back Half and Front Half, with teammate names.
"""

from datetime import date, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from dc_shiftmaster.models import ScheduleSlot, ShiftWindow
from dc_shiftmaster.scheduling import SchedulingEngine

DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
DAY_ABBR = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

# Styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=10)
DATE_FONT = Font(size=9)
MARKER_FONT = Font(bold=True, size=11)
TIME_FONT = Font(size=8, color="444444")
NAME_FONT = Font(size=8, color="444444")
BACK_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FRONT_FILL = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
NIGHT_BACK_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
NIGHT_FRONT_FILL = PatternFill(start_color="D5A6E6", end_color="D5A6E6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _find_cycle_start(year: int) -> date:
    """Find the last Sunday of the previous December.

    The rotation calendar starts from the last Sunday of the prior
    year's December so that weeks run Sunday through Saturday.
    """
    dec31 = date(year - 1, 12, 31)
    # weekday(): Mon=0..Sun=6, Sunday=6
    days_since_sun = (dec31.weekday() - 6) % 7
    return dec31 - timedelta(days=days_since_sun)


def _get_owner_label(engine: SchedulingEngine, d: date) -> str:
    """Return 'B' for back half or 'F' for front half."""
    owner = engine.get_day_owner(d)
    return "B" if owner == "back_half" else "F"


class ExcelExporter:
    """Exports the shift rotation calendar to an Excel spreadsheet.

    Produces a layout matching the ATL Back-Front Shift Rotation format:
    - Weeks run Wednesday through Tuesday
    - Each 2-week block shows dates, day names, B/F markers, and teammate names
    - 4 quarterly columns side by side
    """

    def export(
        self,
        year: int,
        schedule: list[ScheduleSlot],
        engine: SchedulingEngine,
        filepath: str,
        shift_windows: Optional[dict[str, ShiftWindow]] = None,
    ) -> None:
        wb = Workbook()

        # Build slot lookup
        lookup: dict[tuple[str, str], ScheduleSlot] = {}
        for slot in schedule:
            lookup[(slot.date.isoformat(), slot.shift_type)] = slot

        cycle_start = _find_cycle_start(year)
        year_end = date(year, 12, 31)

        # Sheet 1: Merged Calendar
        ws = wb.active
        ws.title = f"{year} Shift Calendar"
        self._write_merged_sheet(ws, year, cycle_start, year_end, engine, lookup, shift_windows)

        # Sheet 2: Coverage Summary
        ws_cov = wb.create_sheet(title="Coverage Summary")
        self._write_coverage_sheet(ws_cov, year, lookup, shift_windows)

        # Sheet 3: Timeline
        ws_timeline = wb.create_sheet(title="Timeline")
        self._write_timeline_sheet(ws_timeline, year, lookup, shift_windows)

        try:
            wb.save(filepath)
        except (OSError, PermissionError) as e:
            raise OSError(f"Cannot save Excel to '{filepath}': {e}") from e

    def _write_sheet(
        self,
        ws,
        year: int,
        cycle_start: date,
        year_end: date,
        engine: SchedulingEngine,
        lookup: dict[tuple[str, str], ScheduleSlot],
        shift_key: str,
        title: str,
        back_fill: PatternFill,
        front_fill: PatternFill,
    ) -> None:
        """Write a full rotation sheet (day or night)."""
        ws.merge_cells("A1:G1")
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT

        row = 3
        current = cycle_start

        while current <= year_end:
            row = self._write_two_week_block(
                ws, row, current, year, engine, lookup,
                shift_key=shift_key, back_fill=back_fill, front_fill=front_fill,
            )
            row += 1
            current += timedelta(days=14)

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _write_two_week_block(
        self,
        ws,
        start_row: int,
        week_start: date,
        year: int,
        engine: SchedulingEngine,
        lookup: dict[tuple[str, str], ScheduleSlot],
        shift_key: str = "day",
        back_fill: PatternFill = BACK_FILL,
        front_fill: PatternFill = FRONT_FILL,
    ) -> int:
        """Write a 2-week block (Week 1 + Week 2) starting from week_start.

        Each block has 4 rows:
          Row 1: Dates (D-Mon format)
          Row 2: Day names (Wednesday, Thursday, ...)
          Row 3: B markers (back half days)
          Row 4: F markers (front half days)
        Then repeats for week 2.

        Returns the next available row.
        """
        row = start_row

        for week_offset in range(2):
            w_start = week_start + timedelta(days=week_offset * 7)

            # Row 1: Dates
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                cell = ws.cell(row=row, column=col_idx + 1)
                if d.year == year or (d.year == year - 1 and d.month == 12):
                    cell.value = f"{d.day}-{d.strftime('%b')}"
                    cell.font = DATE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")

            row += 1

            # Row 2: Day names
            for col_idx, name in enumerate(DAY_NAMES):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = name
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            row += 1

            # Row 3: B markers (back half)
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                if d.year != year and not (d.year == year - 1 and d.month == 12):
                    continue
                owner = _get_owner_label(engine, d)
                cell = ws.cell(row=row, column=col_idx + 1)
                if owner == "B":
                    cell.value = "B"
                    cell.font = MARKER_FONT
                    cell.fill = back_fill
                    cell.alignment = Alignment(horizontal="center")
                    slot = lookup.get((d.isoformat(), shift_key))
                    if slot and slot.teammates != ["nobody"]:
                        names = ", ".join(slot.teammates)
                        cell.value = f"B\n{names}"
                        cell.font = Font(bold=True, size=9)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="top", wrap_text=True
                        )
                cell.border = THIN_BORDER

            row += 1

            # Row 4: F markers (front half)
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                if d.year != year and not (d.year == year - 1 and d.month == 12):
                    continue
                owner = _get_owner_label(engine, d)
                cell = ws.cell(row=row, column=col_idx + 1)
                if owner == "F":
                    cell.value = "F"
                    cell.font = MARKER_FONT
                    cell.fill = front_fill
                    cell.alignment = Alignment(horizontal="center")
                    slot = lookup.get((d.isoformat(), shift_key))
                    if slot and slot.teammates != ["nobody"]:
                        names = ", ".join(slot.teammates)
                        cell.value = f"F\n{names}"
                        cell.font = Font(bold=True, size=9)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="top", wrap_text=True
                        )
                cell.border = THIN_BORDER

            row += 1

        return row

    @staticmethod
    def _calc_end_time(custom_start: str, default_start: str, default_end: str) -> str:
        """Calculate end time for a custom start, preserving the shift duration.

        E.g. default day is 06:00–18:30 = 12.5h. If custom start is 10:00,
        end = 10:00 + 12.5h = 22:30.
        """
        def to_mins(hhmm: str) -> int:
            h, m = hhmm.split(":")
            return int(h) * 60 + int(m)

        def from_mins(m: int) -> str:
            m = m % 1440
            h = m // 60
            mm = m % 60
            return f"{h:02d}:{mm:02d}"

        def_start = to_mins(default_start)
        def_end = to_mins(default_end)
        duration = def_end - def_start
        if duration <= 0:
            duration += 1440

        cust_start = to_mins(custom_start)
        return from_mins(cust_start + duration)

    def _format_cell_content(
        self,
        marker: str,
        shift_window: Optional[ShiftWindow],
        slot: Optional[ScheduleSlot],
    ) -> str:
        """Build multi-line cell text: marker, optional time range, optional names.

        Lines are joined with newlines:
          1. The B/F marker (always present)
          2. "start_time\u2013end_time" (only when shift_window is provided)
          3. Teammate names with individual custom times where applicable

        When a teammate has a custom start time (via slot.teammate_starts),
        their line shows "name (custom_start\u2013custom_end)" instead of just
        the name. Teammates without custom starts are listed normally.
        """
        lines = [marker]

        if shift_window is not None:
            lines.append(f"{shift_window.start_time}\u2013{shift_window.end_time}")

        if slot is not None and slot.teammates != ["nobody"]:
            custom_starts = slot.teammate_starts or {}
            standard_names = []
            custom_lines = []
            for name in slot.teammates:
                custom_start = custom_starts.get(name, "")
                if custom_start and shift_window is not None:
                    custom_end = self._calc_end_time(
                        custom_start, shift_window.start_time, shift_window.end_time
                    )
                    custom_lines.append(f"{name} ({custom_start}\u2013{custom_end})")
                else:
                    standard_names.append(name)
            if standard_names:
                lines.append(", ".join(standard_names))
            for cl in custom_lines:
                lines.append(cl)

        return "\n".join(lines)

    def _write_merged_sheet(
        self,
        ws,
        year: int,
        cycle_start: date,
        year_end: date,
        engine: SchedulingEngine,
        lookup: dict[tuple[str, str], ScheduleSlot],
        shift_windows: Optional[dict[str, ShiftWindow]],
    ) -> None:
        """Write the single merged calendar worksheet."""
        # Title row merged across A1:H1
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{year} Shift Calendar"
        ws["A1"].font = TITLE_FONT

        row = 3
        current = cycle_start

        while current <= year_end:
            row = self._write_merged_two_week_block(
                ws, row, current, year, engine, lookup, shift_windows
            )
            row += 1  # spacing between blocks
            current += timedelta(days=14)

        # Set column widths for columns 1-8 to at least 18 characters
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _write_merged_two_week_block(
        self,
        ws,
        start_row: int,
        week_start: date,
        year: int,
        engine: SchedulingEngine,
        lookup: dict[tuple[str, str], ScheduleSlot],
        shift_windows: Optional[dict[str, ShiftWindow]],
    ) -> int:
        """Write a 2-week block with day+night rows per week. Returns next row.

        Each week has 6 rows:
          Row 1: Dates (D-Mon format) in columns 1-7
          Row 2: Day names (Wednesday..Tuesday) in columns 1-7
          Row 3: Day shift B-markers + times + names (label "Day" in col A, data cols 2-8)
          Row 4: Day shift F-markers + times + names (data cols 2-8)
          Row 5: Night shift B-markers + times + names (label "Night" in col A, data cols 2-8)
          Row 6: Night shift F-markers + times + names (data cols 2-8)
        """
        row = start_row

        day_window = shift_windows.get("day") if shift_windows else None
        night_window = shift_windows.get("night") if shift_windows else None

        for week_offset in range(2):
            w_start = week_start + timedelta(days=week_offset * 7)

            # Row 1: Dates (columns 2-8, aligned with marker columns)
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                cell = ws.cell(row=row, column=col_idx + 2)
                if d.year == year or (d.year == year - 1 and d.month == 12):
                    cell.value = f"{d.day}-{d.strftime('%b')}"
                    cell.font = DATE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")

            row += 1

            # Row 2: Day names (columns 2-8, aligned with marker columns)
            for col_idx, name in enumerate(DAY_NAMES):
                cell = ws.cell(row=row, column=col_idx + 2)
                cell.value = name
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            row += 1

            # Row 3: Day shift B-markers (label in col 1, markers in cols 2-8)
            ws.cell(row=row, column=1).value = "Day"
            ws.cell(row=row, column=1).font = HEADER_FONT
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                if d.year != year and not (d.year == year - 1 and d.month == 12):
                    continue
                owner = _get_owner_label(engine, d)
                cell = ws.cell(row=row, column=col_idx + 2)
                if owner == "B":
                    slot = lookup.get((d.isoformat(), "day"))
                    content = self._format_cell_content("B", day_window, slot)
                    cell.value = content
                    cell.fill = BACK_FILL
                    cell.font = Font(bold=True, size=8)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="top", wrap_text=True
                    )
                cell.border = THIN_BORDER

            row += 1

            # Row 4: Day shift F-markers (markers in cols 2-8)
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                if d.year != year and not (d.year == year - 1 and d.month == 12):
                    continue
                owner = _get_owner_label(engine, d)
                cell = ws.cell(row=row, column=col_idx + 2)
                if owner == "F":
                    slot = lookup.get((d.isoformat(), "day"))
                    content = self._format_cell_content("F", day_window, slot)
                    cell.value = content
                    cell.fill = FRONT_FILL
                    cell.font = Font(bold=True, size=8)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="top", wrap_text=True
                    )
                cell.border = THIN_BORDER

            row += 1

            # Row 5: Night shift B-markers (label in col 1, markers in cols 2-8)
            ws.cell(row=row, column=1).value = "Night"
            ws.cell(row=row, column=1).font = HEADER_FONT
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                if d.year != year and not (d.year == year - 1 and d.month == 12):
                    continue
                owner = _get_owner_label(engine, d)
                cell = ws.cell(row=row, column=col_idx + 2)
                if owner == "B":
                    slot = lookup.get((d.isoformat(), "night"))
                    content = self._format_cell_content("B", night_window, slot)
                    cell.value = content
                    cell.fill = NIGHT_BACK_FILL
                    cell.font = Font(bold=True, size=8)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="top", wrap_text=True
                    )
                cell.border = THIN_BORDER

            row += 1

            # Row 6: Night shift F-markers (markers in cols 2-8)
            for col_idx in range(7):
                d = w_start + timedelta(days=col_idx)
                if d.year != year and not (d.year == year - 1 and d.month == 12):
                    continue
                owner = _get_owner_label(engine, d)
                cell = ws.cell(row=row, column=col_idx + 2)
                if owner == "F":
                    slot = lookup.get((d.isoformat(), "night"))
                    content = self._format_cell_content("F", night_window, slot)
                    cell.value = content
                    cell.fill = NIGHT_FRONT_FILL
                    cell.font = Font(bold=True, size=8)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="top", wrap_text=True
                    )
                cell.border = THIN_BORDER

            row += 1

        return row

    def _write_coverage_sheet(
        self,
        ws,
        year: int,
        lookup: dict[tuple[str, str], ScheduleSlot],
        shift_windows: Optional[dict[str, ShiftWindow]] = None,
    ) -> None:
        """Write a coverage summary with dynamic time blocks from shift settings."""
        import calendar as cal

        # Get shift times (default to 06:00/18:00 if not provided)
        day_start = "06:00"
        day_end = "18:30"
        night_start = "18:00"
        night_end = "06:30"
        if shift_windows:
            if "day" in shift_windows:
                day_start = shift_windows["day"].start_time
                day_end = shift_windows["day"].end_time
            if "night" in shift_windows:
                night_start = shift_windows["night"].start_time
                night_end = shift_windows["night"].end_time

        # Coverage fills by headcount
        cov_0 = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        cov_1 = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cov_2 = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
        cov_3 = PatternFill(start_color="6BAF6B", end_color="6BAF6B", fill_type="solid")

        def _cov_fill(count: int) -> PatternFill:
            if count == 0:
                return cov_0
            elif count == 1:
                return cov_1
            elif count == 2:
                return cov_2
            return cov_3

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Coverage Summary — {year}"
        ws["A1"].font = TITLE_FONT

        # Shift info row
        ws.merge_cells("A2:I2")
        ws["A2"] = f"Day: {day_start}–{day_end}  |  Night: {night_start}–{night_end}"
        ws["A2"].font = Font(size=10, italic=True, color="666666")

        # Dynamic headers based on actual shift times
        headers = [
            "Date", "Day",
            f"Night\n00:00–{day_start}",
            f"Day Shift\n{day_start}–{night_start}",
            f"Overlap AM\n{day_start}–{night_end}",
            f"Overlap PM\n{night_start}–{day_end}",
            f"Night\n{night_start}–24:00",
            "Day Shift\nNames (count)",
            "Night Shift\nNames (count)",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = h
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        num_days = 366 if cal.isleap(year) else 365
        row = 5
        day_abbr = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

        for day_offset in range(num_days):
            d = date(year, 1, 1) + timedelta(days=day_offset)
            date_iso = d.isoformat()

            day_slot = lookup.get((date_iso, "day"))
            night_slot = lookup.get((date_iso, "night"))

            day_names = day_slot.teammates if day_slot else ["nobody"]
            night_names = night_slot.teammates if night_slot else ["nobody"]

            day_count = 0 if day_names == ["nobody"] else len(day_names)
            night_count = 0 if night_names == ["nobody"] else len(night_names)

            am_overlap = day_count + night_count
            pm_overlap = day_count + night_count

            # Date
            c = ws.cell(row=row, column=1)
            c.value = f"{d.day}-{d.strftime('%b')}"
            c.font = DATE_FONT
            c.border = THIN_BORDER

            # Day of week
            c = ws.cell(row=row, column=2)
            c.value = day_abbr[d.weekday()]
            c.font = DATE_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

            # Night early (00:00 to day_start)
            c = ws.cell(row=row, column=3)
            c.value = night_count
            c.fill = _cov_fill(night_count)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

            # Day shift (day_start to night_start)
            c = ws.cell(row=row, column=4)
            c.value = day_count
            c.fill = _cov_fill(day_count)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

            # AM overlap (day_start to night_end)
            c = ws.cell(row=row, column=5)
            c.value = am_overlap
            c.fill = _cov_fill(am_overlap)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

            # PM overlap (night_start to day_end)
            c = ws.cell(row=row, column=6)
            c.value = pm_overlap
            c.fill = _cov_fill(pm_overlap)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

            # Night late (night_start to 24:00)
            c = ws.cell(row=row, column=7)
            c.value = night_count
            c.fill = _cov_fill(night_count)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

            # Day names with count
            c = ws.cell(row=row, column=8)
            if day_names != ["nobody"]:
                c.value = f"({day_count}) {', '.join(day_names)}"
            else:
                c.value = "(0) —"
            c.font = NAME_FONT
            c.border = THIN_BORDER

            # Night names with count
            c = ws.cell(row=row, column=9)
            if night_names != ["nobody"]:
                c.value = f"({night_count}) {', '.join(night_names)}"
            else:
                c.value = "(0) —"
            c.font = NAME_FONT
            c.border = THIN_BORDER

            row += 1

        widths = [12, 8, 14, 14, 14, 14, 14, 35, 35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_timeline_sheet(
        self,
        ws,
        year: int,
        lookup: dict[tuple[str, str], ScheduleSlot],
        shift_windows: Optional[dict[str, ShiftWindow]] = None,
    ) -> None:
        """Write a timeline sheet with 24-hour color bars per day.

        Each row is one day. Columns 3–50 represent 30-minute blocks
        from 00:00 to 24:00 (48 columns). Colored fills show:
          - Orange: day shift active
          - Blue: night shift active
          - Dark green: overlap (both shifts)
          - Red: gap (no coverage)
        """
        import calendar as cal

        # Parse shift times to half-hour indices (0=00:00, 1=00:30, ..., 47=23:30)
        def _time_to_idx(t: str) -> int:
            h, m = int(t.split(":")[0]), int(t.split(":")[1])
            return h * 2 + (1 if m >= 30 else 0)

        day_start_idx = _time_to_idx(shift_windows["day"].start_time) if shift_windows and "day" in shift_windows else 12
        day_end_idx = _time_to_idx(shift_windows["day"].end_time) if shift_windows and "day" in shift_windows else 37
        night_start_idx = _time_to_idx(shift_windows["night"].start_time) if shift_windows and "night" in shift_windows else 36
        night_end_idx = _time_to_idx(shift_windows["night"].end_time) if shift_windows and "night" in shift_windows else 13

        # Fills
        day_fill = PatternFill(start_color="E8943A", end_color="E8943A", fill_type="solid")
        night_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        overlap_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        gap_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        no_fill = PatternFill(fill_type=None)

        # Title
        ws.merge_cells("A1:AZ1")
        ws["A1"] = f"Shift Timeline — {year}"
        ws["A1"].font = TITLE_FONT

        day_s = shift_windows["day"].start_time if shift_windows and "day" in shift_windows else "06:00"
        day_e = shift_windows["day"].end_time if shift_windows and "day" in shift_windows else "18:30"
        night_s = shift_windows["night"].start_time if shift_windows and "night" in shift_windows else "18:00"
        night_e = shift_windows["night"].end_time if shift_windows and "night" in shift_windows else "06:30"
        ws.merge_cells("A2:AZ2")
        ws["A2"] = f"Day: {day_s}–{day_e}  |  Night: {night_s}–{night_e}  |  Orange=Day  Blue=Night  Green=Overlap  Red=Gap"
        ws["A2"].font = Font(size=9, italic=True, color="666666")

        # Hour headers (row 3): columns 3–50 = 48 half-hour blocks
        for i in range(48):
            col = i + 3
            cell = ws.cell(row=3, column=col)
            h = i // 2
            m = "00" if i % 2 == 0 else "30"
            cell.value = f"{h}:{m}" if i % 2 == 0 else ""
            cell.font = Font(size=7)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 3.5

        # Date and Day columns
        ws.cell(row=3, column=1, value="Date").font = HEADER_FONT
        ws.cell(row=3, column=2, value="Day").font = HEADER_FONT
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 6

        num_days = 366 if cal.isleap(year) else 365
        row = 4
        day_abbr = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

        for day_offset in range(num_days):
            d = date(year, 1, 1) + timedelta(days=day_offset)
            date_iso = d.isoformat()

            day_slot = lookup.get((date_iso, "day"))
            night_slot = lookup.get((date_iso, "night"))
            has_day = day_slot and day_slot.teammates != ["nobody"]
            has_night = night_slot and night_slot.teammates != ["nobody"]

            # Date and day-of-week
            ws.cell(row=row, column=1, value=f"{d.day}-{d.strftime('%b')}").font = DATE_FONT
            ws.cell(row=row, column=2, value=day_abbr[d.weekday()]).font = DATE_FONT

            # Build coverage map for each 30-min block
            for i in range(48):
                col = i + 3
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER

                # Determine if day shift covers this block
                if day_start_idx <= day_end_idx:
                    in_day = day_start_idx <= i < day_end_idx
                else:
                    in_day = i >= day_start_idx or i < day_end_idx

                # Determine if night shift covers this block
                # Night typically wraps: e.g. 18:00-06:30 = idx 36..13
                if night_start_idx <= night_end_idx:
                    in_night = night_start_idx <= i < night_end_idx
                else:
                    in_night = i >= night_start_idx or i < night_end_idx

                day_active = in_day and has_day
                night_active = in_night and has_night

                if day_active and night_active:
                    cell.fill = overlap_fill
                elif day_active:
                    cell.fill = day_fill
                elif night_active:
                    cell.fill = night_fill
                elif in_day or in_night:
                    # Shift window exists but nobody assigned = gap
                    cell.fill = gap_fill
                # else: outside all shift windows, leave blank

            row += 1
