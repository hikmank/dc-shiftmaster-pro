"""
Training Rotation Scheduler
- Rotates team members fairly across ATL68 and ATL73 (Mon-Thu)
- Tracks external travel so unavailable people are auto-skipped
- Supports adding/removing team members dynamically
- Persists all state to data.json
"""

import json
import os
import random
from datetime import datetime, timedelta
from collections import deque

DATA_FILE = "data.json"

DEFAULT_EMPLOYEES = [
    "Akeen Bernard", "Anthony Quarles", "Sergio Pirela", "John Fabrizio",
    "Ronnie Baggio", "Aaron Watson", "Jeremiah Bonner", "Achelito Pamphile",
    "Mike Ryan", "Juan Mandorthupp", "Sai Bore", "TeKiya Walls",
    "Adrian Wilder", "Travis Jones", "Connor Scott"
]

LOCATIONS = ["ATL68", "ATL73"]
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday"]
PEOPLE_PER_LOCATION = 2
PEOPLE_PER_DAY = PEOPLE_PER_LOCATION * len(LOCATIONS)  # 4 per day


def default_data():
    return {
        "employees": list(DEFAULT_EMPLOYEES),
        "rotation_queue": list(DEFAULT_EMPLOYEES),
        "external_travel": {},  # { "Name": [{"start": "YYYY-MM-DD", "end": "YYYY-MM-DD", "region": "..."}] }
        "schedules": {},        # { "YYYY-WW": { "Monday": {"ATL68": [...], "ATL73": [...]}, ... } }
        "history": {}           # { "Name": count }
    }


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            data = json.load(f)
        # Ensure all keys exist
        for key, val in default_data().items():
            if key not in data:
                data[key] = val
        return data
    return default_data()


def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)


def get_week_key(date=None):
    """Return ISO year-week string like '2026-W13'."""
    if date is None:
        date = datetime.now()
    iso = date.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def get_week_dates(date=None):
    """Return Mon-Thu dates for the week containing `date`."""
    if date is None:
        date = datetime.now()
    # Find Monday of this week
    monday = date - timedelta(days=date.weekday())
    return [monday + timedelta(days=i) for i in range(4)]  # Mon-Thu


def is_on_external_travel(data, name, date):
    """Check if employee is traveling externally on a given date."""
    date_str = date.strftime("%Y-%m-%d")
    travels = data.get("external_travel", {}).get(name, [])
    for trip in travels:
        if trip["start"] <= date_str <= trip["end"]:
            return True
    return False


def get_available(data, date):
    """Return list of employees available on a given date."""
    return [e for e in data["employees"] if not is_on_external_travel(data, e, date)]


def generate_week_schedule(data, target_date=None):
    """
    Generate a full Mon-Thu schedule for the week of target_date.
    Uses a fair rotation queue: no one repeats until everyone has gone.
    """
    if target_date is None:
        target_date = datetime.now()

    week_key = get_week_key(target_date)
    dates = get_week_dates(target_date)
    schedule = {}
    assigned_this_week = set()

    queue = deque(data["rotation_queue"])

    for i, day_date in enumerate(dates):
        day_name = DAYS[i]
        available = get_available(data, day_date)
        day_assignments = {loc: [] for loc in LOCATIONS}

        needed = PEOPLE_PER_DAY
        chosen = []

        # Pull from front of queue, skip unavailable
        skipped = []
        attempts = 0
        while len(chosen) < needed and attempts < len(queue) + len(skipped) + 1:
            if not queue:
                # Refill queue with everyone not yet chosen this generation
                refill = [e for e in data["employees"] if e not in chosen]
                random.shuffle(refill)
                queue = deque(refill)
                if not queue:
                    break
            candidate = queue.popleft()
            if candidate not in available:
                skipped.append(candidate)
            elif candidate in assigned_this_week and len([e for e in available if e not in assigned_this_week]) >= needed:
                # Try to avoid repeating within same week if possible
                skipped.append(candidate)
            else:
                chosen.append(candidate)
                assigned_this_week.add(candidate)
            attempts += 1

        # Put skipped people back at front of queue
        for s in reversed(skipped):
            queue.appendleft(s)

        # If we still need more, pull from skipped (unavoidable repeats)
        while len(chosen) < needed and skipped:
            fallback = skipped.pop(0)
            if fallback in available:
                chosen.append(fallback)
                try:
                    queue.remove(fallback)
                except ValueError:
                    pass

        # Distribute chosen across locations
        for j, loc in enumerate(LOCATIONS):
            start = j * PEOPLE_PER_LOCATION
            end = start + PEOPLE_PER_LOCATION
            day_assignments[loc] = chosen[start:end]

        schedule[day_name] = day_assignments

        # Update history
        for person in chosen:
            data["history"][person] = data["history"].get(person, 0) + 1

    data["rotation_queue"] = list(queue)
    data["schedules"][week_key] = schedule
    save_data(data)
    return week_key, schedule


def add_employee(data, name):
    name = name.strip()
    if name and name not in data["employees"]:
        data["employees"].append(name)
        data["rotation_queue"].append(name)
        save_data(data)
        return True
    return False


def remove_employee(data, name):
    if name in data["employees"]:
        data["employees"].remove(name)
        if name in data["rotation_queue"]:
            data["rotation_queue"].remove(name)
        save_data(data)
        return True
    return False


def add_external_travel(data, name, start_date, end_date, region=""):
    if name not in data["employees"]:
        return False
    if name not in data["external_travel"]:
        data["external_travel"][name] = []
    data["external_travel"][name].append({
        "start": start_date,
        "end": end_date,
        "region": region
    })
    save_data(data)
    return True


def remove_external_travel(data, name, index):
    if name in data["external_travel"] and 0 <= index < len(data["external_travel"][name]):
        data["external_travel"][name].pop(index)
        save_data(data)
        return True
    return False


def import_travel_from_excel(data, filepath):
    """
    Import travel entries from an Excel file.
    Expected columns (case-insensitive, flexible matching):
      - Employee / Name
      - Start / Start Date
      - End / End Date
      - Region / Notes / Location (optional)

    Returns (imported_count, skipped, errors)
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # Read header row and map columns
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value or "").strip().lower())

    col_map = {"name": None, "start": None, "end": None, "region": None}
    for i, h in enumerate(headers):
        if h in ("employee", "name", "employee name", "team member"):
            col_map["name"] = i
        elif h in ("start", "start date", "start_date", "from", "travel start"):
            col_map["start"] = i
        elif h in ("end", "end date", "end_date", "to", "travel end"):
            col_map["end"] = i
        elif h in ("region", "notes", "location", "destination", "city"):
            col_map["region"] = i

    if col_map["name"] is None or col_map["start"] is None or col_map["end"] is None:
        wb.close()
        return 0, 0, ["Could not find required columns. Need: Employee/Name, Start/Start Date, End/End Date"]

    imported = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=2):
        try:
            raw_name = row[col_map["name"]].value
            raw_start = row[col_map["start"]].value
            raw_end = row[col_map["end"]].value
            raw_region = row[col_map["region"]].value if col_map["region"] is not None else ""

            if not raw_name or not raw_start or not raw_end:
                skipped += 1
                continue

            name = str(raw_name).strip()

            # Handle date objects or strings
            if hasattr(raw_start, "strftime"):
                start_str = raw_start.strftime("%Y-%m-%d")
            else:
                start_str = str(raw_start).strip()[:10]

            if hasattr(raw_end, "strftime"):
                end_str = raw_end.strftime("%Y-%m-%d")
            else:
                end_str = str(raw_end).strip()[:10]

            region = str(raw_region or "").strip()

            # Fuzzy match employee name (case-insensitive)
            matched_name = None
            for emp in data["employees"]:
                if emp.lower() == name.lower():
                    matched_name = emp
                    break
            if not matched_name:
                # Try partial match
                for emp in data["employees"]:
                    if name.lower() in emp.lower() or emp.lower() in name.lower():
                        matched_name = emp
                        break

            if not matched_name:
                errors.append(f"Employee '{name}' not found on team — skipped")
                skipped += 1
                continue

            # Validate dates
            datetime.strptime(start_str, "%Y-%m-%d")
            datetime.strptime(end_str, "%Y-%m-%d")

            add_external_travel(data, matched_name, start_str, end_str, region)
            imported += 1

        except Exception as e:
            errors.append(f"Row error: {e}")
            skipped += 1

    wb.close()
    return imported, skipped, errors


def export_schedule_csv(data, week_key, filepath):
    """Export a week's schedule to CSV."""
    import csv
    schedule = data.get("schedules", {}).get(week_key)
    if not schedule:
        return False
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Day"] + LOCATIONS)
        for day in DAYS:
            day_data = schedule.get(day, {})
            row = [day]
            for loc in LOCATIONS:
                people = day_data.get(loc, [])
                row.append(", ".join(people))
            writer.writerow(row)
    return True


def export_schedule_excel(data, week_key, filepath):
    """Export a week's schedule to Excel."""
    from openpyxl import Workbook
    schedule = data.get("schedules", {}).get(week_key)
    if not schedule:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = week_key
    ws.append(["Day"] + LOCATIONS)
    for day in DAYS:
        day_data = schedule.get(day, {})
        row = [day]
        for loc in LOCATIONS:
            people = day_data.get(loc, [])
            row.append(", ".join(people))
        ws.append(row)
    wb.save(filepath)
    return True


def export_travel_csv(data, filepath):
    """Export all travel entries to CSV."""
    import csv
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Employee", "Start Date", "End Date", "Region"])
        for name in sorted(data.get("external_travel", {}).keys()):
            for trip in data["external_travel"][name]:
                writer.writerow([name, trip["start"], trip["end"], trip.get("region", "")])
    return True


def export_travel_excel(data, filepath):
    """Export all travel entries to Excel."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Travel"
    ws.append(["Employee", "Start Date", "End Date", "Region"])
    for name in sorted(data.get("external_travel", {}).keys()):
        for trip in data["external_travel"][name]:
            ws.append([name, trip["start"], trip["end"], trip.get("region", "")])
    wb.save(filepath)
    return True


def export_history_csv(data, filepath):
    """Export assignment history to CSV."""
    import csv
    history = data.get("history", {})
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Employee", "Total Assignments", "In Queue"])
        queue = data.get("rotation_queue", [])
        for name in sorted(data["employees"]):
            count = history.get(name, 0)
            in_queue = "Yes" if name in queue else "No"
            writer.writerow([name, count, in_queue])
    return True


def export_history_excel(data, filepath):
    """Export assignment history to Excel."""
    from openpyxl import Workbook
    history = data.get("history", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "History"
    ws.append(["Employee", "Total Assignments", "In Queue"])
    queue = data.get("rotation_queue", [])
    for name in sorted(data["employees"]):
        count = history.get(name, 0)
        in_queue = "Yes" if name in queue else "No"
        ws.append([name, count, in_queue])
    wb.save(filepath)
    return True
